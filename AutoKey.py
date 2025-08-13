import openpyxl
import pyautogui
import time
from pynput import keyboard

# Global flag to stop the program
stop_flag = False

def read_excel_column(file_path, sheet_name, start_cell):
    """Read values from a column in an Excel file until an empty cell is encountered."""
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Determine the starting row and column
    start_column = openpyxl.utils.cell.column_index_from_string(start_cell[:1])  # Column letter to number
    start_row = int(start_cell[1:])  # Starting row number

    # Read values from the column
    values = []
    for row in range(start_row, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=start_column).value
        if cell_value is None:  # Stop at the first empty cell
            break
        values.append(cell_value)

    return values

def type_to_program(data):
    """Type data directly into another program."""
    # Type the data as a string
    pyautogui.typewrite(str(data))  # Utilizes pyautogui library to simulate user typing
    pyautogui.press("enter")  # Will simulate the enter key being pressed
    pyautogui.press("esc")  # Dismisses the popup confirmation after keying in a quantity

def on_press(key):
    """Handle key press events."""
    global stop_flag
    try:
        if key.char == 'q':  # Stop the program when 'q' is pressed
            pyautogui.press("esc")
            stop_flag = True
            print("Cancel signal received. Stopping program...")
            return False  # Stop the listener
    except AttributeError:
        pass

def main():
    global stop_flag

    # Configuration
    print("IMPORTANT: Please Make Sure You have the PI Program OPEN and READY with the correct Location before beginning.")
    print("...")
    print("To stop the program, press 'q' at any time.")
    print("===================================")
    excel_file = (input("What is your Excel File Path: "))[1:-1]
    print("...")
    sheet_name = input("What is your Sheet Named (Ex. Sheet1): ")
    print("...")
    start_cell = input("What is your Starting Cell (Ex. C2): ")
    print("...")
    
    try:
        # Step 1: Read data from the Excel column
        values = read_excel_column(excel_file, sheet_name, start_cell)
        if not values:
            print("No data found in the specified column.")
            return

        print("Generating number list...")
        print("...")
        print(f"Values to type: {values}")
        print("...")
        print("PROGRAM WILL BEGIN SHORTLY")
        print("///")
        print("Please select the beginning box in the Physical Inventory Program now.")
        time.sleep(5)
        print("Beginning keying in 3 Seconds...")
        time.sleep(1)
        print("3")
        time.sleep(1)
        print("2")
        time.sleep(1)
        print("1")
        print("===================================")
        print("Values Typed")
        print("===================================")

        # Start listening for key presses
        listener = keyboard.Listener(on_press=on_press)
        listener.start()

        # Step 2: Automate typing each value into the program
        for value in values:
            if stop_flag:
                print("Process stopped by user.")
                break
            print(f"Typing value: {value}")
            type_to_program(value)
            time.sleep(0.6)  # Optional: Adjust time delay between inputs if needed

        # Stop the listener
        listener.stop()
        if not stop_flag:
            print("===================================")
            print("All data typed successfully.")
        print("===================================")
        while True:
            user_input = input("Enter 'exit' to close the program: ")
            if user_input.lower() == 'exit':
                print("Exiting program...")
                break
        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    main()
