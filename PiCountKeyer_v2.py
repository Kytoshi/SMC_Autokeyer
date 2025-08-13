import tkinter as tk
import customtkinter as ctk
import openpyxl
from tkinter import *
from tkinter import filedialog
from PIL import Image, ImageTk
import pyautogui
import time
import threading
from pynput import keyboard

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

##### Program Global Settings #####

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

global stop_flag, entryVar, labelVar, cellVar, pathVar, sheetVar, validExcel
stop_flag = False
validExcel = []

##### Clear Page Configuration #####
""" Page for the clearing function input """

class ClearPage(ctk.CTkFrame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller

        global entryVar, labelVar

        entryVar = tk.StringVar()
        labelVar = tk.StringVar()

        self.configure(fg_color="#384959")

        title2 = ctk.CTkLabel(self,
        text="C L E A R", 
        font=("Inter", 26, "bold"), 
        fg_color="#6A89A7",
        text_color="white", 
        anchor="center", 
        width=200,
        height=50
        )
        title2.place(relx=0.5, rely=0.2, anchor=CENTER)

        # BACK BUTTON ---

        back_default_img = ImageTk.PhotoImage(Image.open("components/defaultback.png").resize((30, 30), Image.Resampling.BICUBIC))
        
        self.back_button = ctk.CTkButton(self, 
            image=back_default_img, 
            width=20,
            height=20,
            text="",
            fg_color="transparent",
            hover_color="#6A89A7",
            command=lambda: self.backhome()
            )
        self.back_button.place(relx=0.1, rely=0.2, anchor=CENTER)

        prompt2 = ctk.CTkLabel(self,
        text="How many boxes do you need to clear?",
        font=("Inter", 15, "bold"),
        text_color="white",
        anchor="center",
        width=200
        )
        prompt2.place(relx=0.5, rely=0.35, anchor=CENTER)

        promptDesc = ctk.CTkLabel(self,
        text="Note: A 5 Second Countdown will start after pressing Confirm. \n Please ensure you are ready to select the PI Count Box...",
        font=("Inter", 13, "italic"),
        text_color="white",
        anchor="center",
        width=200
        )
        promptDesc.place(relx=0.5, rely=0.42, anchor=CENTER)

        self.clear_entry = ctk.CTkEntry(self, 
            placeholder_text="Enter number of boxes...", 
            width=300, 
            height=40, 
            border_color="#6A89A7", 
            textvariable=entryVar)
        self.clear_entry.place(relx=0.5, rely=0.55, anchor=CENTER)

        begin_clearing_button = ctk.CTkButton(self, 
            text="Confirm", 
            font=("Inter", 17, "bold"), 
            width=200, 
            height=40,
            fg_color="#88BDF2",
            text_color="black", 
            command=self.validate_input)
        begin_clearing_button.place(relx=0.5, rely=0.75, anchor=CENTER)

    def validate_input(self):
        value = entryVar.get()
        try:
            if int(value) > 0:
                self.clear_entry.configure(border_color="#6A89A7")  # Reset to default gray
                self.controller.show_page("countDown")
            else:
                raise ValueError
        except ValueError:
            self.clear_entry.configure(border_color="red")  # Show red border
    
    def backhome(self):
            """Return to Home and reset flags."""
            entryVar.set("")  # Clear the entry field
            self.clear_entry.configure(border_color="#6A89A7")  # Reset to default gray
            self.controller.show_page("Home")  # Go back to home

##### Count Down Page Configuration #####
""" Supporter Page for the Clear Page."""

class countDownPage(ctk.CTkFrame):
    def __init__(self, parent, controller, countdown_time=5):
        super().__init__(parent)
        self.controller = controller
        self.countdown_time = countdown_time
        self._cancelled = False  # Add cancellation flag

        self.configure(fg_color="#384959")

        self.count_text = ctk.CTkLabel(self, 
        text="Starting in", 
        font=("Inter", 26, "bold"), 
        text_color="White")
        self.count_text.place(relx=0.5, rely=0.3, anchor=CENTER)

        self.count = ctk.CTkLabel(self, 
            text="", 
            font=("Inter", 40, "bold"), 
            text_color="white")
        self.count.place(relx=0.5, rely=0.4, anchor=CENTER)

        # Home button (hidden at start)
        self.home_button = ctk.CTkButton(
            self,
            text="Return to Home",
            font=("Inter", 17, "bold"),
            fg_color="#88BDF2",
            text_color="black",
            width=200,
            height=40,
            command=lambda: self.return_home()  # Use method to return to home
        )
        self.home_button.place(relx=0.5, rely=0.6, anchor=CENTER)
        self.home_button.lower()  # Hide initially

        # Cancel button
        self.cancel_button = ctk.CTkButton(
            self,
            text="Cancel",
            font=("Inter", 17, "bold"),
            fg_color="#88BDF2",
            text_color="black",
            width=200,
            height=40,
            command=self.cancel_countdown  # Cancel countdown
        )
        self.cancel_button.place(relx=0.5, rely=0.7, anchor=CENTER)

    def cancel_countdown(self):
        """Cancel countdown and reset the flag."""
        self._cancelled = True  # Set the cancel flag
        global stop_flag
        stop_flag = True  # Set stop_flag to True to stop the process
        self.count.configure(text="Cancelled", text_color="red")
        entryVar.set("")  # Clear the entry field after processing
        self.count_text.configure(text="")  # Clear the "Starting in" text
        self.home_button.lift()  # Show the Home button so user can go back manually
        self.cancel_button.lower()  # Hide the Cancel button after cancellation

    def return_home(self):
        """Return to Home and reset flags."""
        global stop_flag
        stop_flag = False  # Reset stop flag for future runs
        self._cancelled = False  # Reset cancel flag
        self.count_text.configure(text="Starting in", text_color="white")
        self.controller.show_page("Home")  # Go back to home

    def start_countdown(self):
        self.controller.pages["Clear"].clear_entry.configure(border_color="#6A89A7")
        self.remaining = self.countdown_time
        self.cancel_button.lift()
        self.home_button.lower()  # Hide button if revisiting
        self.count.configure(text="", text_color="white")
        self._update_timer()

    def _update_timer(self):
        if self.remaining >= 0 and not self._cancelled:  # Check if not cancelled
            self.count.configure(text=self.remaining)
            self.remaining -= 1
            self.update()  # Force UI update to display Cancel button
            self.after(1000, self._update_timer)
        else:
            if not self._cancelled:
                self.count_text.configure(text="")
                self.count.configure(text="Clearing has started!")
                self.after(500, self.on_clear_finished)  # Call on_clear_finished after countdown
            else:
                # If cancelled, stop further action and reset
                self.count.configure(text="Cancelled", text_color="red")

    def on_clear_finished(self):
        """Start the clearing process in a separate thread."""
        thread = threading.Thread(target=self.start_clear_thread)
        thread.daemon = True # Make thread a daemon so it exits when the main program exits
        thread.start()

    def start_clear_thread(self):
        """Runs the clearing logic and updates UI from main thread."""
        self.start_clear()

    def start_clear(self):
        """Converts the entry value into an integer and starts the clearing process; cancels process if stop_flag is set."""
        global entryVar, stop_flag, labelVar
        boxes = entryVar.get()
        try:
            boxes = int(boxes)
        except ValueError:
            self.count.configure(text="Invalid input", text_color="red")
            return

        labelVar.set(boxes)
        entryVar.set("")  # Clear the entry field after processing

        for count in range(boxes):
            if stop_flag:
                print("Process stopped by user.")
                break
            self.clear_boxes(-1)
            time.sleep(0.1)  # Optional: Adjust time delay between inputs if needed

        if not stop_flag:
            self.after(1000, self._update_done)  # Update UI after process is done
        else:
            self.after(1000, self._update_cancelled)  # Update UI after process is cancelled

    def _update_done(self):
        """Update the UI after the clearing process is done."""
        self.count.configure(text="Done!", text_color="green")
        self.cancel_button.lower()  # Hide the cancel button after finishing
        self.home_button.lift()  # Show the home button after finishing

    def _update_cancelled(self):
        """Update the UI after the clearing process is cancelled."""
        self.count.configure(text="Cancelled", text_color="red")
        self.home_button.lift()  # Show the home button after cancelling

    def clear_boxes(self, data):
        """Deleting Count in Input."""
        # Type the data as a string
        pyautogui.typewrite(str(data))
        time.sleep(0.1)
        pyautogui.press("enter")  # Press Enter to press confirm button
        time.sleep(0.1)
        pyautogui.press("esc")  # Press esc to dismiss confirm dialog box

##### Key Page Configuration #####
""" Main Function to Fill in the inputs."""

class KeyPage(ctk.CTkFrame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller

        global pathVar, cellVar, sheetVar

        pathVar = tk.StringVar()
        cellVar = tk.StringVar()
        sheetVar = tk.StringVar()

        self.configure(fg_color="#384959")

        self.title2 = ctk.CTkLabel(self,
            text="K E Y", 
            font=("Inter", 26, "bold"), 
            text_color="white",
            fg_color="#6A89A7",
            anchor="center",
            width=200,
            height=50
            )
        self.title2.place(relx=0.5, rely=0.2, anchor=CENTER)

        self.back_default_img = ImageTk.PhotoImage(Image.open("components/defaultback.png").resize((30, 30), Image.Resampling.BICUBIC))
        
        self.key_page_button = ctk.CTkButton(self, 
            image=self.back_default_img, 

            width=20, height=20,
            text="",
            fg_color="transparent",
            hover_color="#6A89A7",
            command=lambda: self.backhome()
            )
        self.key_page_button.place(relx=0.1, rely=0.2, anchor=CENTER)

        promptDesc = ctk.CTkLabel(self,
        text="Note: A 5 Second Countdown will start after pressing Confirm. \n Please ensure you are ready to select the PI Count Box...",
        font=("Inter", 13, "italic"),
        text_color="white",
        anchor="center",
        width=200
        )
        promptDesc.place(relx=0.5, rely=0.31, anchor=CENTER)

        self.file_path_text = ctk.CTkLabel(self, 
            text="Excel File Path:", 
            font=("Inter", 15, "bold"), 
            text_color="white")
        self.file_path_text.place(relx=0.3, rely=0.4, anchor=CENTER)

        self.file_path_entry = ctk.CTkEntry(self, 
            placeholder_text="Enter Excel File Path...", 
            width=300, 
            height=40, 
            textvariable=pathVar
            )
        self.file_path_entry.pack(padx=(160, 3), side=ctk.LEFT)

        self.browse_button = ctk.CTkButton(self, 
            text="Browse...", 
            width=100, 
            height=40,
            font=("Inter", 12, "bold"),
            fg_color="#88BDF2",
            text_color="black",
            command=self.browse_file
        )
        self.browse_button.pack(padx=(0, 0), side=ctk.LEFT)

        self.startingcell_text = ctk.CTkLabel(self, 
            text="Starting Cell:", 
            font=("Inter", 15, "bold"), 
            text_color="white")
        self.startingcell_text.place(relx=0.285, rely=0.62, anchor=CENTER)

        self.cell_entry = ctk.CTkEntry(self,
            placeholder_text="Enter Starting Cell (Ex. C2)", 
            width=55, 
            height=40, 
            textvariable=cellVar
            )
        self.cell_entry.place(relx=0.4, rely=0.62, anchor=CENTER)

        self.sheet_text = ctk.CTkLabel(self, 
            text="Sheet Name:", 
            font=("Inter", 15, "bold"), 
            text_color="white")
        self.sheet_text.place(relx=0.56, rely=0.62, anchor=CENTER)

        self.sheet_entry = ctk.CTkEntry(self,
            placeholder_text="Enter Sheet Name (Ex. Sheet1)", 
            width=100, 
            height=40, 
            textvariable=sheetVar
            )
        self.sheet_entry.place(relx=0.71, rely=0.62, anchor=CENTER)

        self.begin_keying_button = ctk.CTkButton(self,
            text="Confirm", 
            font=("Inter", 17, "bold"), 
            width=200, 
            height=40,
            fg_color="#88BDF2",
            text_color="black", 
            command=lambda: self.validate_input()
        )
        self.begin_keying_button.place(relx=0.5, rely=0.8, anchor=CENTER)

        self.error_text = ctk.CTkLabel(self, text="No Data Found", font=("Inter", 15, "bold"), text_color="red")
        self.error_text.place(relx=0.5, rely=0.71, anchor=CENTER)
        self.error_text.lower()

    def backhome(self):
        """Return to Home and reset flags."""
        cellVar.set("")
        pathVar.set("")
        sheetVar.set("")
        self.error_text.configure(text="VALIDATING DATA...", text_color="white")
        self.error_text.lower()
        self.controller.show_page("Home")  # Go back to home

    def validate_input(self):
        global cellVar, sheetVar, pathVar, validExcel

        self.error_text.configure(text="VALIDATING DATA...", text_color="white")
        self.error_text.lift()
        self.update()

        # Clean up file path (remove any "" if found)
        excelfile = pathVar.get().replace('"', '')
        pathVar.set(excelfile)  # Optional: update entry with cleaned path

        cell = cellVar.get()
        sheet = sheetVar.get()

        # Initialize the container if not already
        if 'validExcel' not in globals():
            validExcel = tk.StringVar()

        try:
            excelValue = self.read_excel_column(excelfile, sheet, cell)

            # Check if excelValue is empty or None
            if not excelValue or excelValue == ["", None]:
                self.error_text.configure(text="NO DATA FOUND.", text_color="red")
                self.error_text.lift()
                return  # stop execution if invalid

            # Check if all values are valid (non-negative integers)
            invalid_values = [value for value in excelValue if not isinstance(value, int) or value < 0]
            if invalid_values:
                self.error_text.configure(text="NOT VALID ENTRY", text_color="red")
                self.error_text.lift()
                return  # stop execution if invalid

            # If validation passed
            self.error_text.lower()
            validExcel = excelValue  # Store the validated data
            self.controller.show_page("countDown2")

            # Clear inputs AFTER validation
            cellVar.set("")
            pathVar.set("")
            sheetVar.set("")

        except Exception as e:
            print("Validation Error:", e)
            self.error_text.configure(text=str(e), text_color="red")
            self.error_text.lift()
            
    def read_excel_column(self, file_path, sheet_name, start_cell):
        """Read values from a column in an Excel file until an empty cell is encountered."""
        try:
            print(f"Reading from: {file_path}")
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            print(f"Available sheets: {workbook.sheetnames}")
            
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found.")
            
            sheet = workbook[sheet_name]

            from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
            col_letter, row = coordinate_from_string(start_cell)
            start_column = column_index_from_string(col_letter)
            start_row = row

            # Read values from the column
            values = []
            for r in range(start_row, sheet.max_row + 1):
                cell_value = sheet.cell(row=r, column=start_column).value
                if cell_value is None:
                    break
                values.append(cell_value)

            print("Values found:", values)
            return values  # <--- Important: return the list of values

        except Exception as e:
            print(f"Error reading Excel file: {e}")
            return []

    def browse_file(self):
        """Opens file dialog and sets the selected path into the entry."""
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if file_path:
            pathVar.set(file_path)

class countDownPage2(ctk.CTkFrame):
    def __init__(self, parent, controller, countdown_time=5):
        super().__init__(parent)
        self.controller = controller
        self.countdown_time = countdown_time
        self._cancelled = False  # Add cancellation flag

        self.configure(fg_color="#384959")

        self.count_text = ctk.CTkLabel(self, 
            text="Starting in", 
            font=("Inter", 26, "bold"), 
            text_color="white")
        self.count_text.place(relx=0.5, rely=0.3, anchor=CENTER)

        self.count = ctk.CTkLabel(self, 
            text="", 
            font=("Inter", 40, "bold"), 
            text_color="#213448")
        self.count.place(relx=0.5, rely=0.4, anchor=CENTER)

        # Home button (hidden at start)
        self.home_button = ctk.CTkButton(
            self,
            text="Return to Home",
            font=("Inter", 17, "bold"),
            width=200,
            height=40,
            command=lambda: self.return_home()  # Use method to return to home
        )
        self.home_button.place(relx=0.5, rely=0.6, anchor=CENTER)
        self.home_button.lower()  # Hide initially

        # Cancel button
        self.cancel_button = ctk.CTkButton(
            self,
            text="Cancel",
            font=("Inter", 17, "bold"),
            fg_color="#88BDF2",
            text_color="black",
            width=200,
            height=40,
            command=self.cancel_countdown  # Cancel countdown
        )
        self.cancel_button.place(relx=0.5, rely=0.7, anchor=CENTER)

    def cancel_countdown(self):
        """Cancel countdown and reset the flag."""
        self._cancelled = True  # Set the cancel flag
        global stop_flag
        stop_flag = True  # Set stop_flag to True to stop the process
        self.count.configure(text="Cancelled", text_color="red")
        cellVar.set("")  # Clear the entry field after processing
        pathVar.set("")
        sheetVar.set("")
        self.count_text.configure(text="")  # Clear the "Starting in" text
        self.home_button.lift()  # Show the Home button so user can go back manually
        self.cancel_button.lower()  # Hide the Cancel button after cancellation

    def return_home(self):
        """Return to Home and reset flags."""
        global stop_flag
        stop_flag = False  # Reset stop flag for future runs
        self._cancelled = False  # Reset cancel flag
        self.count_text.configure(text="Starting in", text_color="white")
        self.controller.show_page("Home")  # Go back to home

    def start_countdown(self):
        self.controller.pages["Key"].cell_entry.configure(border_color="#6A89A7")
        self.controller.pages["Key"].sheet_entry.configure(border_color="#6A89A7")
        self.controller.pages["Key"].file_path_entry.configure(border_color="#6A89A7")
        self.remaining = self.countdown_time
        self.cancel_button.lift()
        self.home_button.lower()  # Hide button if revisiting
        self.count.configure(text="", text_color="white")
        self._update_timer()

    def _update_timer(self):
        if self.remaining >= 0 and not self._cancelled:  # Check if not cancelled
            self.count.configure(text=self.remaining)
            self.remaining -= 1
            self.update()  # Force UI update to display Cancel button
            self.after(1000, self._update_timer)
        else:
            if not self._cancelled:
                self.count_text.configure(text="")
                self.count.configure(text="Keying has started!")
                self.count.configure(text_color="white")
                self.after(500, self.on_key_finished)  # Call on_key_finished after countdown
            else:
                # If cancelled, stop further action and reset
                self.count.configure(text="Cancelled", text_color="red")

    def on_key_finished(self):
        """Start the keying process in a separate thread."""
        thread = threading.Thread(target=self.start_key_thread)
        thread.daemon = True # Make thread a daemon so it exits when the main program exits
        thread.start()

    def start_key_thread(self):
        """Runs the clearing logic and updates UI from main thread."""
        self.start_key()

    def start_key(self):
        global stop_flag, validExcel
        values = validExcel
        validExcel = []
        for value in values:
            if stop_flag:
                print("Process stopped by user.")
                break
            print(f"Typing value: {value}")
            self.type_to_program(value)
            time.sleep(0.6)
        
        if not stop_flag:
            validExcel = []
            self.after(1000, self._update_done)  # Update UI after process is done
        else:
            validExcel = []
            self.after(1000, self._update_cancelled)  # Update UI after process is cancelled

        
    def _update_done(self):
        """Update the UI after the clearing process is done."""
        self.count.configure(text="Done!", text_color="green")
        self.cancel_button.lower()  # Hide the cancel button after finishing
        self.home_button.lift()  # Show the home button after finishing

    def _update_cancelled(self):
        """Update the UI after the clearing process is cancelled."""
        self.count.configure(text="Cancelled", text_color="red")
        self.home_button.lift()  # Show the home button after cancelling

    def type_to_program(self, data):
        """Type data directly into another program."""
        # Type the data as a string
        pyautogui.typewrite(str(data))  # Utilizes pyautogui library to simulate user typing
        pyautogui.press("enter")  # Will simulate the enter key being pressed
        pyautogui.press("esc")  # Dismisses the popup confirmation after keying in a quantity


##### Home Page Configuration #####
""" Hub Page to access the different Functions. """

class HomePage(ctk.CTkFrame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller

        self.configure(fg_color="#384959")

        title1 = ctk.CTkLabel(self,
        text="P I   C O U N T   A U T O K E Y", 
        font=("Inter", 24, "bold"),
        text_color="white",
        bg_color="#6A89A7", 
        anchor="center", 
        width=450,
        height= 70
        )
        title1.place(relx=0.5, rely=0.2, anchor=CENTER)

        subtitle1 = ctk.CTkLabel(self,
        text="Are You Keying or Clearing?",
        font=("Inter", 15, "bold"),
        text_color="white",
        anchor="center",
        width=700
        )
        subtitle1.place(relx=0.5, rely=0.33, anchor=CENTER)

        buttons_frame = ctk.CTkFrame(self, fg_color="#384959", width=400, height=60)
        buttons_frame.place(relx=0.5, rely=0.5, anchor=CENTER)

        key_page_button = ctk.CTkButton(buttons_frame, 
            text="KEY", 
            font=("Inter", 17, "bold"), 
            width=100, 
            height=50,
            fg_color="#88BDF2",
            text_color="black",
            command=lambda: controller.show_page("Key"))
        key_page_button.pack(padx=(0,33), side=ctk.LEFT)
        
        back_button = ctk.CTkButton(buttons_frame, 
            text="CLEAR", 
            font=("Inter", 17, "bold"), 
            width=100, 
            height=50,
            fg_color="#88BDF2",
            text_color="black",
            command=lambda: controller.show_page("Clear"))
        back_button.pack(padx=(0,0), side=ctk.RIGHT)

##### Program Set up #####
""" Sets up Window for Program, keep track of the pages for navigation. """

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("PI Count AutoKey")
        self.geometry("720x480")

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # Container for all pages
        container = ctk.CTkFrame(self)
        container.pack(fill="both", expand=True)

        # Let container expand
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        # Dictionary to hold references to pages
        self.pages = {}

       # Create and grid pages, hide all initially
        for PageClass, name in [(HomePage, "Home"), 
            (KeyPage, "Key"), 
            (ClearPage, "Clear"), 
            (countDownPage, "countDown"),
            (countDownPage2, "countDown2")
            ]:
            page = PageClass(container, self)
            self.pages[name] = page
            page.grid(row=0, column=0, sticky="nsew")
            page.lower()  # Hide by default

        self.show_page("Home")  # Show the home page first

    def show_page(self, page_name):
        """Show a page by name."""
        for page in self.pages.values():
            page.lower()

        # Raise only the one we want
        self.pages[page_name].tkraise()

        # Start Countdown if on the countdown page
        if page_name == "countDown":
            self.pages[page_name].start_countdown()
        
        if page_name == "countDown2":
            self.pages[page_name].start_countdown()

if __name__ == "__main__":
    app = App()
    app.mainloop()
